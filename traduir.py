from openpyxl import load_workbook
from deep_translator import GoogleTranslator

wb = load_workbook('fitxer.xlsx')
f = open ('configuracio.txt', 'r')
linies = f.readlines()
wb = load_workbook(linies[4].split('=')[1].strip())
fulls = {}
for idioma in linies[0].split('=')[1].split(','):
#for idioma in linies[0].split(','):
	wb.create_sheet(idioma.strip())
	fulls[idioma.strip()] = wb[idioma.strip()]
#wb.create_sheet('arabe')
#ws = wb["Català"]
ws = wb[linies[5].split('=')[1].strip()]
#ws2 = wb['castellano']
#ws3 = wb['arabe']
fila_i = linies[2].split('=')[1].split('-')[0]
fila_f = linies[2].split('=')[1].split('-')[1]
colu_i = linies[3].split('=')[1].split('-')[0]
colu_f = linies[3].split('=')[1].split('-')[1]

for i in range(int(fila_i),int(fila_f)+1):
	for j in range (int(colu_i),int(colu_f)+1):
		print ('cel·la: ' + str(i) + ',' + str(j))
		valor = ws.cell(row=i, column=j)
		if not valor.value == None:
			for full in fulls:
				fulls[full].cell(row=i, column=j, value=GoogleTranslator(source=linies[1].split('=')[1].strip(), target=full).translate(valor.value))
wb.save(linies[4].split('=')[1].strip())
wb.close()
